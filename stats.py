import csv
import re
import os
import string
import pdfkit
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:
    """Класс для представления датасета

    Attributes:
        file_name (str): Путь к файлу
        vacancies_objects (list[Vacancy]): Список вакансий
    """
    def __init__(self, file_name, vacancies_objects):
        """Инициализирует датасет

        Args:
            file_name (str): Путь к файлу
            vacancies_objects (list[Vacancy]): Список вакансий
        """
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects


class Vacancy:
    """
    Вакансия

    Attributes:
        name (str): Название вакансии
        salary (Salary): Заработная плата
        area_name (str): Название города
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        """
        Args:
            name (str): Название вакансии
            salary (Salary): Заработная плата
            area_name (str): Название города
            published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    """Класс для представления зарплаты

    Attributes:
        salary_from (float): Нижняя граница вилки оклада
        salary_to (float): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Инициализирует объект Salary

        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency


class DataSetReader:
    """
    Класс, используемый для считывания датасета из csv-файла

    Attributes:
        path (str): Путь к файлу
    """
    def __init__(self, path):
        """
        Инициализирует объект DataSetReader

        Args:
            path (str): Путь к файлу
        """
        self.path = path

    def read(self):
        """
        Считывает датасет из файла

        Returns:
             Dataset: датасет
        """
        csv = self._read_csv()
        lines = self._parse_csv(*csv)
        return self._convert_to_dataset(lines)

    def _read_csv(self):
        """
        Считывает строки из csv-файла

        Returns:
             tuple[list[str], list[list[str]]]: заголовок и список строк таблицы
        """
        with open(self.path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            try:
                header = reader.__next__()
            except StopIteration:
                raise Exception("Пустой файл")
            lines = []
            for line in reader:
                lines.append(line)
            return header, lines

    def _parse_csv(self, header, lines):
        """
        Очищает файл от пустых строк и html-тегов и преобразует индивидуальные строки в словари

        Returns:
            list[dict[str, *]]: Список строк таблицы
        """
        lines_clean = []
        for line in lines:
            if '' in line or len(line) != len(header):
                continue
            d = {}
            for i, item in enumerate(line):
                item = item.replace('\r', '').split('\n')
                for j, sub in enumerate(item):
                    item[j] = re.sub(' +', ' ',
                                     re.sub('<.*?>', '', sub.replace('\u2002', ' ').replace('\u00a0', ' '))).strip()
                if len(item) > 1:
                    d[header[i]] = '\n'.join(item)
                else:
                    d[header[i]] = item[0]
            lines_clean.append(d)
        return lines_clean

    def _convert_to_dataset(self, lines):
        """
        Преобразует строки таблицы в датасет

        Args:
            lines (list[dict[str, *]]): Список строк таблицы
        """
        vacancies = []
        for l in lines:
            salary = Salary(l['salary_from'], l['salary_to'], l['salary_currency'])
            vacancy = Vacancy(l['name'], salary, l['area_name'], l['published_at'])
            vacancies.append(vacancy)
        return DataSet(self.path, vacancies)


class MeanRepresentation:
    """Класс для представления среднего арифметического

    Attributes:
        sum (int): Сумма элементов
        count (int): Количество элементов
    """
    def __init__(self):
        self.sum = 0
        self.count = 0

    def add(self, num):
        """
        Добавляет новый элемент
        """
        self.sum += num
        self.count += 1

    def mean(self):
        """
        Возвращает значение среднего арифметического

        Returns:
            float: Среднее арифметическое
        """
        if self.count == 0:
            return 0
        return self.sum / self.count

    def __int__(self):
        return int(self.mean())

    def __str__(self):
        return str(int(self))

    def __repr__(self) -> str:
        return str(self)


class MyDefaultDict(defaultdict):
    """Перезаписывает метод __repr__ defaultdict на стандартный"""
    def __str__(self) -> str:
        return dict.__repr__(self)


class Statistics:
    """Класс, содержащий статистику по вакансиям

    Attributes:
        dataset (DataSet): датасет
        profession_name (str): Название профессии, по которой генерируется дополнительная статистика
        salary_dynamic (dict): Динамика уровня зарплат по годам
        count_dynamic (dict): Динамика количества вакансий по годам
        salary_dynamic_filtered (dict): Динамика уровня зарплат по годам для выбранной профессии
        count_dynamic_filtered (dict): Динамика количества вакансий по годам для выбранной профессии
        city_salaries (dict): Уровень зарплат по городам (в порядке убывания)
        city_counts (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, dataset, profession_name):
        """Инициализирует объект Statistics и генерирует статистику по вакансиям

        Args:
            dataset (DataSet): датасет
            profession_name (str): Название профессии, по которой генерируется дополнительная статистика
        """
        self.dataset = dataset
        self.profession_name = profession_name
        self._generate()

    def _generate(self):
        """
        Генерирует статистику по вакансиям
        """
        name = self.profession_name
        dataset = self.dataset
        defaultdict = MyDefaultDict

        salary_dynamic = defaultdict(lambda: MeanRepresentation())
        salary_dynamic_filtered = defaultdict(lambda: MeanRepresentation())
        count_dynamic = defaultdict(lambda: 0)
        count_dynamic_filtered = defaultdict(lambda: 0)
        city_counts = defaultdict(lambda: 0)
        city_salaries = defaultdict(lambda: MeanRepresentation())
        one_percent = int(len(dataset.vacancies_objects) / 100)

        for v in dataset.vacancies_objects:
            year = int(v.published_at[:4])
            salary = int(
                (float(v.salary.salary_to) + float(v.salary.salary_from)) / 2 * currency_to_rub[
                    v.salary.salary_currency])
            count_dynamic[year] += 1
            salary_dynamic[year].add(salary)
            if name in v.name:
                count_dynamic_filtered[year] += 1
                salary_dynamic_filtered[year].add(salary)
            city_counts[v.area_name] += 1
            city_salaries[v.area_name].add(salary)

        allowed_cities = []
        for city in city_counts:
            if city_counts[city] >= one_percent:
                allowed_cities.append(city)
        city_count_list = list(zip(city_counts.keys(), city_counts.values()))
        city_count_list.sort(key=lambda x: x[1], reverse=True)
        city_count_list = city_count_list[:10]
        for i in range(len(city_count_list)):
            city_count_list[i] = city_count_list[i][0], round(city_count_list[i][1] / len(dataset.vacancies_objects), 4)
        city_counts = dict(filter(lambda x: x[0] in allowed_cities, city_count_list))

        city_salaries_list = list(
            filter(lambda x: x[0] in allowed_cities, zip(city_salaries.keys(), city_salaries.values())))
        city_salaries_list.sort(key=lambda x: x[1].mean(), reverse=True)
        city_salaries_list = city_salaries_list[:10]
        city_salaries = dict(city_salaries_list)

        if len(count_dynamic_filtered) == 0:
            count_dynamic_filtered = {2022: 0}
        if len(salary_dynamic_filtered) == 0:
            salary_dynamic_filtered = {2022: MeanRepresentation()}

        self.salary_dynamic = salary_dynamic
        self.count_dynamic = count_dynamic
        self.salary_dynamic_filtered = salary_dynamic_filtered
        self.count_dynamic_filtered = count_dynamic_filtered
        self.city_salaries = city_salaries
        self.city_counts = city_counts

    def print(self):
        """
        Выводит статистику по вакансиям
        """
        print('Динамика уровня зарплат по годам:', self.salary_dynamic)
        print('Динамика количества вакансий по годам:', self.count_dynamic)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.salary_dynamic_filtered)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.count_dynamic_filtered)
        print('Уровень зарплат по городам (в порядке убывания):', self.city_salaries)
        print('Доля вакансий по городам (в порядке убывания):', self.city_counts)

    @property
    def years(self):
        return self.salary_dynamic.keys()


class Report:
    """Класс для генерации отчета
    """
    def __init__(self, stats: Statistics):
        """Инициализирует объект

        Args:
            stats (Statistics): статистика
        """
        self.stats = stats

    def generate_excel(self, filename='report.xlsx'):
        """Генерирует отчет в формате excel"""
        wb = Workbook()
        self._generate_excel_year_statistic(wb.create_sheet('Статистика по годам'))
        self._generate_excel_city_statistic(wb.create_sheet('Статистика по городам'))
        del wb['Sheet']
        wb.save(filename)

    def _generate_excel_year_statistic(self, ws):
        """Генерирует страницу со статистикой по годам"""
        ws.append(('Год', 'Средняя зарплата', f'Средняя зарплата - {self.stats.profession_name}',
                   'Количество вакансий', f'Количество вакансий - {self.stats.profession_name}'))
        self._excel_make_bold(ws, ws.calculate_dimension())
        for year in self.stats.years:
            ws.append((year, int(self.stats.salary_dynamic[year]), int(self.stats.salary_dynamic_filtered[year]),
                       self.stats.count_dynamic[year], self.stats.count_dynamic_filtered[year]))

        self._excel_make_border(ws, ws.calculate_dimension())
        self._excel_set_column_widths(ws)

    def _generate_excel_city_statistic(self, ws):
        """Генерирует страницу со статистикой по городам"""
        ws.append(('Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'))
        self._excel_make_bold(ws, ws.calculate_dimension())
        col = 1
        row = 2
        for city, salary in self.stats.city_salaries.items():
            ws.cell(row, col, city)
            ws.cell(row, col + 1, int(salary))
            row += 1
        col = 4
        row = 2
        for city, count in self.stats.city_counts.items():
            ws.cell(row, col, city)
            ws.cell(row, col + 1, count).number_format = FORMAT_PERCENTAGE_00
            row += 1

        self._excel_make_border(ws, 'A1:B11')
        self._excel_make_border(ws, 'D1:E11')
        self._excel_set_column_widths(ws)

    def _excel_make_bold(self, ws, _range):
        for row in ws[_range]:
            for cell in row:
                cell.font = Font(bold=True)

    def _excel_make_border(self, ws, _range):
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        for row in ws[_range]:
            for cell in row:
                cell.border = border

    def _excel_set_column_widths(self, ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate_image(self, filename='graph.png'):
        """Генерирует изображение с графиками"""
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        self._make_plt1(ax1)
        self._make_plt2(ax2)
        self._make_plt3(ax3)
        self._make_plt4(ax4)

        plt.tight_layout()
        fig.savefig(filename)

    def _make_plt1(self, ax):
        """диаграмма - уровень зарплат по годам для вывода динамики уровня зарплат по годам как общий, так и для выбранной профессии"""
        xticks = np.arange(len(self.stats.years)), self.stats.years
        width = 0.4
        x1 = list(map(int, self.stats.salary_dynamic.values()))
        x2 = list(map(int, self.stats.salary_dynamic_filtered.values()))
        max_salary = max(max(x1), max(x2))

        ax.bar(xticks[0] - width / 2, x1, width=width,
                label='средняя з/п')
        ax.bar(xticks[0] + width / 2, x2, width=width,
                label=f'з/п {self.stats.profession_name.lower()}')
        ax.set_xticks(*xticks, rotation=90, fontsize=8)
        ax.set_yticks(np.arange(0, max_salary, 20000), fontsize=8)
        ax.set_title('Уровень зарплат по годам')
        ax.grid(axis='y')
        ax.legend(fontsize=8)

    def _make_plt2(self, ax):
        """диаграмма - количество вакансий по годам как общий, так и для выбранной профессии"""
        xticks = np.arange(len(self.stats.years)), self.stats.years
        width = 0.4
        x1 = list(map(int, self.stats.count_dynamic.values()))
        x2 = list(map(int, self.stats.count_dynamic_filtered.values()))
        max_count = max(max(x1), max(x2))

        ax.bar(xticks[0] - width / 2, x1, width=width,
               label='Количество вакансий')
        ax.bar(xticks[0] + width / 2, x2, width=width,
               label=f'Количество вакансий\n{self.stats.profession_name.lower()}')
        ax.set_xticks(*xticks, rotation=90, fontsize=8)
        ax.set_yticks(np.arange(0, max_count, 20000), fontsize=8)
        ax.set_title('Количество вакансий по годам')
        ax.grid(axis='y')
        ax.legend(fontsize=8)

    def _make_plt3(self, ax):
        """горизонтальная диаграмма - уровень зарплат по городам"""
        keys = list(map(lambda x: x.replace('-', '-\n').replace(' ', '\n'), self.stats.city_salaries.keys()))
        yticks = np.arange(len(keys)), keys
        y = list(map(int, self.stats.city_salaries.values()))
        max_salary = max(y)

        ax.barh(yticks[0], y)
        ax.set_yticks(*yticks, fontsize=6, ha='right', va='center')
        ax.set_xticks(np.arange(0, max_salary, 20000), fontsize=8)
        ax.set_title('Уровень зарплат по городам')
        ax.invert_yaxis()
        ax.grid(axis='x')

    def _make_plt4(self, ax):
        """круговая диаграмма - количество вакансий по городам"""
        labels = list(self.stats.city_counts.keys())
        labels.append('Другие')
        x = list(self.stats.city_counts.values())
        x.append(1 - sum(x))

        ax.pie(x, labels=labels, textprops={'fontsize': 6})
        ax.set_title('Доля вакансий по городам')

    def generate_pdf(self, filename='report.pdf'):
        """Генерирует отчет в формате pdf, содержащий таблицу и графики"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template.html")
        html = template.render(stats=self.stats, os=os, int=int)
        pdfkit.from_string(html, filename, options={"enable-local-file-access": None})


def main():
    """
    Входная точка программы.
    Получает на вход путь к файлу и название профессии, генерирует по ним статистику и сохраняет отчет во всех
    возможных форматах.
    """
    file = input('Введите название файла: ')
    name = input('Введите название профессии: ')
    dataset = DataSetReader(file).read()

    stats = Statistics(dataset, name)
    stats.print()
    report = Report(stats)
    report.generate_excel()
    report.generate_image()
    report.generate_pdf()


if __name__ == '__main__':
    main()
